from openpyxl import *
import os, os.path
from  datetime import  date
from cal import Months

class TimebookSetup:

    def Save(self):
        directory_to = 'timesheets'
        file_name = self.get_year() + '_timeheets.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        wb.save(os.path.join(directory_to, file_name))

    ## there should always be a template file in the templates directory ##
    def locate_template(self, directory):
        program_dir = directory    
        template_path = program_dir + '/templates/timesheet_template.xlsx'
        return template_path    


    def locate_timebook(self, directory):
        program_dir = directory
        file_name = self.get_year() + '_timeheets.xlsx'
        file_path = program_dir + '/timesheets/' + file_name
        return file_path

    def get_day(self):
        today = date.today()
        day_num = today.day
        #month_name = Months[str(month_num)]
        return str(day_num)

    def get_month(self):
        today = date.today()
        month_num = today.month
        month_name = Months[str(month_num)]
        return str(month_name)


    def get_year(self):
        today = date.today()
        year_num = today.year    
        return str(year_num)


    def create_timebook(self, program_dir):
        directory_to = 'timesheets'
        file_name = self.get_year() + '_timeheets.xlsx'
        if not os.path.isdir(directory_to):
            os.makedirs(directory_to)
        tb_exists = os.path.exists(program_dir + '/timesheets/' + file_name )
        if tb_exists == True:
            ## Prompt if file already exists , override? ##
            print('Timebook for ' + self.get_year() + ' already exists')
            prompt= input('Override? Y/N :')
            if prompt in ['Y', 'y']:
                temp_wb.save(os.path.join(directory_to, file_name)) 
                print('Timebook for ' + self.get_year() + ' created')
            else:
                print()
                pass     
        else:
            temp_wb.save(os.path.join(directory_to, file_name)) 
            print('Timebook for ' + self.get_year() + ' created')          

            
    def create_timeheet(self):
        if self.get_month() in wb.sheetnames:
            pass
        else:     
            print('Sheet ' + self.get_month() + ' added in workbook')   
            wb.create_sheet(self.get_month())
            self.Save()


class Cloner:       

    def rename_Sheet1(self):
        sheet = wb['Sheet1']
        sheet.title = 'Template' 


    def copy_from_template(self):
        temp_sheet = wb['Template']
        active_sheet = wb[TimebookSetup.get_month()]
        for i in range(1, 100):
            for j in range(1, 26):
                active_sheet.cell(row=i,column=j).value = temp_sheet.cell(row=i,column=j).value

    def Clone(self):
        self.rename_Sheet1()
        self.copy_from_template()
        print('Template Cloned')



if __name__ == "__main__":
        
    ## user required to insert path to program ##
    program_dir = 'C:/Users/Bernard/Dropbox/Personal/python/auto_timesheet'

    Setup = TimebookSetup()
    Cloner = Cloner()
    
    ## template workbook object ##
    temp_wb = load_workbook(str(Setup.locate_template(program_dir)))

    ## current year workbook object ##
    wb = load_workbook(str(Setup.locate_timebook(program_dir)))
    
    ## active sheet set as current month ##
    active_sheet = wb[TimebookSetup.get_month()]


### add to class ###
    ## following appends work descriptions to excel file
    weeks = 4
    days = 5 
    
    active_sheet = wb[TimebookSetup.get_month()]    
    
    week = 1
    for w in range(weeks):
        temp = []
        temp.clear()
        temp.append('week' + str(week))
        active_sheet.append(temp)
        week += 1
        for i in range(days):
            desc_value = input('Description: ')
            day_list = [(str(TimebookSetup.get_day()), desc_value, '*')] 
            for d in day_list:                           
                active_sheet.append(d)   

    
    
    Setup.Save()
